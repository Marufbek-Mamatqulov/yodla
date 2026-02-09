import openpyxl
import json
import re
from collections import Counter

wb = openpyxl.load_workbook("SO'Z BOYLIGI (A1, A2, B1, B2, C1, C2).xlsx")

all_words = []

# ============================================================
# Helper: parse "word — translation" using em-dash (U+2014)
# ============================================================
def split_word_translation(text):
    """Split 'english — uzbek' into (english, uzbek)."""
    if not text:
        return None, None
    text = str(text).strip()
    # Try em-dash first (most common)
    if '\u2014' in text:
        parts = text.split('\u2014', 1)
        return parts[0].strip(), parts[1].strip() if len(parts) > 1 else ''
    # Try en-dash
    if '\u2013' in text:
        parts = text.split('\u2013', 1)
        return parts[0].strip(), parts[1].strip() if len(parts) > 1 else ''
    # Try " - " (space-hyphen-space)
    if ' - ' in text:
        parts = text.split(' - ', 1)
        return parts[0].strip(), parts[1].strip() if len(parts) > 1 else ''
    return text.strip(), ''


def is_header_row(text):
    """Check if a row is a section/unit header."""
    if not text:
        return True
    t = str(text).strip().upper()
    if re.match(r'^UNIT\s+\d', t):
        return True
    headers = ['VOCABULARY FROM', 'TOPIC VOCABULARY', 'WORD PATTERNS',
               'WORD FORMATION', 'PREPOSITIONAL PHRASES', 'PHRASES AND',
               'USE OF ENGLISH', 'CONFUSING', 'PHRASAL VERB']
    for h in headers:
        if t.startswith(h):
            return True
    if '#TOPIC' in t or '#WORD' in t or '#PHRASAL' in t or '#PREPOSITIONAL' in t:
        return True
    return False


# ============================================================
# 1. A1–C2 level sheets
#    Col A = number, Col B = "word — translation"
# ============================================================
for level in ['A1', 'A2', 'B1', 'B2', 'C1', 'C2']:
    if level not in wb.sheetnames:
        continue
    ws = wb[level]
    data_col = 1  # Column B (0-indexed)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        cell = row[data_col] if len(row) > data_col and row[data_col] else None
        if not cell:
            continue
        eng, uzb = split_word_translation(cell)
        if eng and not is_header_row(eng):
            all_words.append({
                'english': eng,
                'uzbek': uzb,
                'level': level,
                'category': 'level'
            })

print(f"A1-C2 words: {len(all_words)}")

# ============================================================
# 2. Fellar 650 ta (Verbs)
#    Col A = "word — translation"
# ============================================================
verbs_start = len(all_words)
if 'Fellar 650 ta' in wb.sheetnames:
    ws = wb['Fellar 650 ta']
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        cell = row[0] if row[0] else None
        if not cell:
            continue
        eng, uzb = split_word_translation(cell)
        if eng and not is_header_row(eng):
            all_words.append({
                'english': eng,
                'uzbek': uzb,
                'level': 'Verbs',
                'category': 'verbs'
            })
print(f"Verbs: {len(all_words) - verbs_start}")

# ============================================================
# 3. Destination B1
#    Col A = "word – translation"
# ============================================================
dest_b1_start = len(all_words)
current_topic = ''
if 'Destination B1 ozbekcha tarjima' in wb.sheetnames:
    ws = wb['Destination B1 ozbekcha tarjima']
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        cell = str(row[0]).strip() if row[0] else ''
        if not cell:
            continue
        if is_header_row(cell):
            current_topic = cell
            continue
        eng, uzb = split_word_translation(cell)
        if eng:
            all_words.append({
                'english': eng,
                'uzbek': uzb,
                'level': 'Dest B1',
                'category': 'dest_b1',
                'topic': current_topic
            })
print(f"Destination B1: {len(all_words) - dest_b1_start}")

# ============================================================
# 4. Destination B2
#    Col A = "word - translation"
# ============================================================
dest_b2_start = len(all_words)
current_topic = ''
if 'Destination B2 ozbekcha tarjima' in wb.sheetnames:
    ws = wb['Destination B2 ozbekcha tarjima']
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        cell = str(row[0]).strip() if row[0] else ''
        if not cell:
            continue
        upper = cell.upper()
        if '#' in upper:
            current_topic = cell
            continue
        eng, uzb = split_word_translation(cell)
        if eng:
            all_words.append({
                'english': eng,
                'uzbek': uzb,
                'level': 'Dest B2',
                'category': 'dest_b2',
                'topic': current_topic
            })
print(f"Destination B2: {len(all_words) - dest_b2_start}")

# ============================================================
# 5. Destination C1&C2
#    Col A = English, Col B = Uzbek
# ============================================================
dest_c_start = len(all_words)
if 'Destination C1&C2 ozbekcha tarj' in wb.sheetnames:
    ws = wb['Destination C1&C2 ozbekcha tarj']
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        eng = str(row[0]).strip() if row[0] else ''
        uzb = str(row[1]).strip() if len(row) > 1 and row[1] else ''
        if not eng or is_header_row(eng):
            continue
        all_words.append({
            'english': eng,
            'uzbek': uzb,
            'level': 'Dest C1C2',
            'category': 'dest_c1c2'
        })
print(f"Destination C1&C2: {len(all_words) - dest_c_start}")

# ============================================================
# 6. Phrasal Verbs
#    Col A = phrasal verb, Col B = Uzbek, Col C = English definition
# ============================================================
pv_start = len(all_words)
if 'Phrasal verbs Destination B2,C1' in wb.sheetnames:
    ws = wb['Phrasal verbs Destination B2,C1']
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        eng = str(row[0]).strip() if row[0] else ''
        uzb = str(row[1]).strip() if len(row) > 1 and row[1] else ''
        defn = str(row[2]).strip() if len(row) > 2 and row[2] else ''
        if not eng or is_header_row(eng):
            continue
        uzb = uzb.replace('\n', ' ').replace('\r', ' ').strip()
        defn = defn.replace('\n', ' ').replace('\r', ' ').strip()
        all_words.append({
            'english': eng,
            'uzbek': uzb,
            'level': 'Phrasal',
            'category': 'phrasal',
            'definition': defn
        })
print(f"Phrasal Verbs: {len(all_words) - pv_start}")

# ============================================================
# Build stats & export
# ============================================================
level_counts = Counter()
cat_counts = Counter()
for w in all_words:
    level_counts[w['level']] += 1
    cat_counts[w['category']] += 1

data = {
    'words': all_words,
    'stats': dict(sorted(level_counts.items())),
    'categoryCounts': dict(sorted(cat_counts.items())),
    'totalWords': len(all_words)
}

with open('data.json', 'w', encoding='utf-8') as f:
    json.dump(data, f, ensure_ascii=False, indent=None)

print(f"\n=== TOTAL: {len(all_words)} words ===")
print(f"Stats by level: {dict(sorted(level_counts.items()))}")
print(f"Stats by category: {dict(sorted(cat_counts.items()))}")
print("Done! data.json created.")
